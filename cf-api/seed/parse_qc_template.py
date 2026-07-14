# -*- coding: utf-8 -*-
# ============================================================
# parse_qc_template.py — อ่านแท็บ "แบบฟอร์มตรวจ QC" จากไฟล์ต้นแบบ
#   (บ้านคุณวริษฐา) → รายการเกณฑ์ QC มาตรฐาน → JSON + SQL seed
#
# ผลลัพธ์:
#   seed/qc_criteria.json      — array ของเกณฑ์ (ไว้ใช้ในหน้าเว็บ/ตรวจสอบ)
#   seed/qc_criteria_seed.sql  — INSERT OR REPLACE ลงตาราง qc_criteria (D1)
#
# ★ แหล่งความจริง = ไฟล์ต้นแบบ (BLUEPRINT §1). ฟอร์มจริงมี 26 ข้อ หมวด A–I
#   (SESSION-3.md เขียนสรุปไว้ 24 ข้อ A–H — ยึดของจริงจากไฟล์)
#
# วิธีรัน (Windows console พิมพ์ไทยไม่ได้ → เขียนลงไฟล์ UTF-8 อย่างเดียว):
#   python cf-api/seed/parse_qc_template.py "<path ไฟล์ xlsx>"
# ============================================================
import sys
import os
import re
import json

try:
    import openpyxl
except ImportError:
    sys.exit("ต้องติดตั้งก่อน: pip install openpyxl")

DEFAULT_XLSX = r"C:\Users\User\Downloads\DSTR - สำเนาของ QC-Overall-Dashboard-บ้านคุณวริษฐา-แก้ไข-28-6-69.xlsx"
FORM_SHEET_INDEX = 0  # แท็บแรก = "แบบฟอร์มตรวจ QC"

# แถวหัวตาราง = 8 (ลำดับ | รายการตรวจสอบ | เกณฑ์ | วิธี | Defect | ผลตรวจ | หมายเหตุ)
# รายการเริ่มแถว 9 (section header) ถึงก่อน "สรุปผลการตรวจสอบ"
COL_SEQ, COL_ITEM, COL_ACCEPT, COL_METHOD, COL_DEFECT = 1, 2, 3, 4, 5

SECTION_RE = re.compile(r"^([A-Z])\.\s*(.+)$")          # "A. งานวัสดุปิดผิว ..."
SEQ_RE = re.compile(r"^\d+(\.\d+)?$")                    # "1.0", "12.0"
DEFECT_CLASS_RE = re.compile(r"^\s*\[(C|M|Mn)\]")        # [C] / [M] / [Mn]


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return "" if v is None else str(v).strip()


def parse(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[FORM_SHEET_INDEX]

    criteria = []
    cur_section = ""
    cur_section_name = ""

    for r in range(9, ws.max_row + 1):
        a = cell(ws, r, COL_SEQ)
        b = cell(ws, r, COL_ITEM)

        # เจอบล็อกสรุป → จบรายการ
        if a.startswith("สรุปผลการตรวจสอบ") or a.startswith("Summary"):
            break

        # section header: มีเฉพาะคอลัมน์ A รูปแบบ "X. ชื่อหมวด"
        m = SECTION_RE.match(a)
        if m and not b:
            cur_section = m.group(1)
            cur_section_name = m.group(2).strip()
            continue

        # item: คอลัมน์ A เป็นเลขลำดับ + มีชื่อรายการในคอลัมน์ B
        if SEQ_RE.match(a) and b:
            seq = float(a)
            defect_text = cell(ws, r, COL_DEFECT)
            dcm = DEFECT_CLASS_RE.match(defect_text)
            defect_class = dcm.group(1) if dcm else ""
            criteria.append({
                "criteria_id": "QC-%02d" % int(round(seq)),
                "section": cur_section,
                "section_name": cur_section_name,
                "seq": seq,
                "item": b,
                "acceptance": cell(ws, r, COL_ACCEPT),
                "method": cell(ws, r, COL_METHOD),
                "defects": defect_text,
                "defect_class": defect_class,
                "active": "TRUE",
            })

    return criteria


def sql_escape(s):
    return s.replace("'", "''")


def to_sql(criteria):
    lines = [
        "-- qc_criteria seed — สร้างอัตโนมัติจาก parse_qc_template.py (ห้ามแก้มือ)",
        "-- ใช้: wrangler d1 execute dstr-db --file=seed/qc_criteria_seed.sql  [--local|--remote]",
        "-- idempotent: INSERT OR REPLACE ตาม criteria_id",
        "",
    ]
    for c in criteria:
        lines.append(
            "INSERT OR REPLACE INTO qc_criteria "
            "(criteria_id, section, section_name, seq, item, acceptance, method, defects, defect_class, active) "
            "VALUES ('{cid}', '{sec}', '{secname}', {seq}, '{item}', '{acc}', '{method}', '{def}', '{dclass}', 'TRUE');".format(
                cid=sql_escape(c["criteria_id"]),
                sec=sql_escape(c["section"]),
                secname=sql_escape(c["section_name"]),
                seq=c["seq"],  # ตัวเลข ไม่ต้อง quote
                item=sql_escape(c["item"]),
                acc=sql_escape(c["acceptance"]),
                method=sql_escape(c["method"]),
                dclass=sql_escape(c["defect_class"]),
                **{"def": sql_escape(c["defects"])}
            )
        )
    return "\n".join(lines) + "\n"


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx_path):
        sys.exit("ไม่พบไฟล์: " + xlsx_path)

    criteria = parse(xlsx_path)
    here = os.path.dirname(os.path.abspath(__file__))

    json_path = os.path.join(here, "qc_criteria.json")
    sql_path = os.path.join(here, "qc_criteria_seed.sql")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(criteria, f, ensure_ascii=False, indent=2)
    with open(sql_path, "w", encoding="utf-8") as f:
        f.write(to_sql(criteria))

    # print เฉพาะ ASCII (Windows console cp874 พิมพ์ไทยไม่ได้)
    sections = {}
    for c in criteria:
        sections.setdefault(c["section"], 0)
        sections[c["section"]] += 1
    print("parsed criteria:", len(criteria))
    print("sections:", ", ".join("%s=%d" % (k, sections[k]) for k in sorted(sections)))
    classes = {}
    for c in criteria:
        classes[c["defect_class"] or "?"] = classes.get(c["defect_class"] or "?", 0) + 1
    print("defect_class:", classes)
    print("wrote:", os.path.basename(json_path), "+", os.path.basename(sql_path))


if __name__ == "__main__":
    main()
